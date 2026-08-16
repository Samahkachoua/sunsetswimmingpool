import csv
import io
import json
import math
import os
import re
import time
from collections import defaultdict, deque
from datetime import date, datetime, timedelta
from datetime import date as date_cls
from fastapi import APIRouter, Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from postgrest.exceptions import APIError

from app.supabase_client import get_admin_user, sign_in_admin, supabase

app = FastAPI(title="Sunset Swimming Pool")

app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")

# Falls back to the production domain if unset, so canonical URLs are still
# correct if a deployment forgets to set it — set SITE_URL=http://localhost:8000
# (or similar) in .env for local dev so canonical tags don't point at prod.
SITE_URL = os.environ.get("SITE_URL", "https://sunsetswimmingpool.com").rstrip("/")
templates.env.globals["SITE_URL"] = SITE_URL

SESSION_COOKIE = "sb_access_token"

PAGE_SIZE_OPTIONS = [25, 50, 100, 200, 500]
DEFAULT_PAGE_SIZE = 25


def _short_date(value) -> str:
    if not value:
        return "—"
    parsed = value if isinstance(value, date) else date.fromisoformat(str(value)[:10])
    return parsed.strftime("%b %d, %Y")


def _time12(value: str) -> str:
    """Format a "HH:MM" or "HH:MM:SS" string as 12-hour time, e.g. "9:00 AM"."""
    if not value:
        return "—"
    hour, minute = (int(part) for part in value.split(":")[:2])
    period = "AM" if hour < 12 else "PM"
    hour_12 = hour % 12 or 12
    return f"{hour_12}:{minute:02d} {period}"


def _parse_list_params(
    request: Request,
    sortable_columns: dict,
    default_sort: str,
    default_dir: str = "asc",
    sort_cookie: str | None = None,
    page_size_cookie: str | None = None,
    page_cookie: str | None = None,
) -> tuple[int, int, str, str]:
    """Parse page/sort_by/sort_dir/page_size, all from cookies.

    This state is kept out of the URL (like the filter cookies) so interacting
    with a list doesn't litter the address bar with query params.
    """
    page_state = _get_filters_cookie(request, page_cookie) if page_cookie else {}
    try:
        page = max(1, int(page_state.get("page", 1)))
    except (TypeError, ValueError):
        page = 1

    page_size_state = _get_filters_cookie(request, page_size_cookie) if page_size_cookie else {}
    page_size = page_size_state.get("page_size", DEFAULT_PAGE_SIZE)
    if page_size not in PAGE_SIZE_OPTIONS:
        page_size = DEFAULT_PAGE_SIZE

    sort_state = _get_filters_cookie(request, sort_cookie) if sort_cookie else {}

    sort_by = sort_state.get("sort_by", default_sort)
    if sort_by not in sortable_columns:
        sort_by = default_sort

    sort_dir = sort_state.get("sort_dir", default_dir)
    if sort_dir not in ("asc", "desc"):
        sort_dir = default_dir

    return page, page_size, sort_by, sort_dir


def _fetch_page(build_query, page: int, page_size: int):
    """build_query(start, end) -> executed response with count='exact'. Clamps out-of-range pages.

    PostgREST errors (PGRST103) instead of returning empty data when the requested
    offset is beyond the available rows, so an out-of-range page has to be probed
    for the real total first, then re-fetched at the clamped page.
    """
    def run(p: int):
        start = (p - 1) * page_size
        end = start + page_size - 1
        return build_query(start, end)

    try:
        result = run(page)
    except APIError as exc:
        if exc.code != "PGRST103":
            raise
        total = run(1).count or 0
        page = max(1, math.ceil(total / page_size)) if total else 1
        result = run(page)

    total = result.count or 0
    total_pages = max(1, math.ceil(total / page_size)) if total else 1
    if page > total_pages:
        page = total_pages
        result = run(page)
        total = result.count or 0
    return result.data, total, page, total_pages


def _export_response(fmt: str, title: str, headers: list[str], rows: list[list], filename_base: str):
    """Build a CSV or Excel download with a title row above the header row."""
    if fmt == "csv":
        text_buffer = io.StringIO()
        writer = csv.writer(text_buffer)
        writer.writerow([title])
        writer.writerow(headers)
        writer.writerows(rows)
        byte_buffer = io.BytesIO(text_buffer.getvalue().encode("utf-8-sig"))
        return StreamingResponse(
            byte_buffer,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.csv"'},
        )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Export"
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    worksheet.append(headers)
    for cell in worksheet[2]:
        cell.font = Font(bold=True)
    for row in rows:
        worksheet.append(row)
    for col_index, header in enumerate(headers, start=1):
        values = [header] + [row[col_index - 1] for row in rows]
        width = min(max((len(str(v)) for v in values if v is not None), default=10) + 2, 40)
        worksheet.column_dimensions[get_column_letter(col_index)].width = width

    byte_buffer = io.BytesIO()
    workbook.save(byte_buffer)
    byte_buffer.seek(0)
    return StreamingResponse(
        byte_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'},
    )


def _age_bounds_to_dob_range(age_min: int | None, age_max: int | None):
    """Convert an inclusive age range into a (min_dob, max_dob) date_of_birth range."""
    today = date.today()

    def shift_years(d: date, years: int) -> date:
        try:
            return d.replace(year=d.year - years)
        except ValueError:  # Feb 29 shifted onto a non-leap year
            return d.replace(month=2, day=28, year=d.year - years)

    max_dob = shift_years(today, age_min) if age_min is not None else None
    min_dob = shift_years(today, age_max + 1) + timedelta(days=1) if age_max is not None else None
    return min_dob, max_dob


ENROLLMENTS_FILTER_COOKIE = "enrollments_filters"
PARTICIPANTS_FILTER_COOKIE = "participants_filters"
PAYMENTS_FILTER_COOKIE = "payments_filters"
RESERVATIONS_FILTER_COOKIE = "reservations_filters"

CYCLES_SORT_COOKIE = "cycles_sort"
PARTICIPANTS_SORT_COOKIE = "participants_sort"
ENROLLMENTS_SORT_COOKIE = "enrollments_sort"
PAYMENTS_SORT_COOKIE = "payments_sort"

CYCLES_PAGE_SIZE_COOKIE = "cycles_page_size"
PARTICIPANTS_PAGE_SIZE_COOKIE = "participants_page_size"
ENROLLMENTS_PAGE_SIZE_COOKIE = "enrollments_page_size"
PAYMENTS_PAGE_SIZE_COOKIE = "payments_page_size"

CYCLES_PAGE_COOKIE = "cycles_page"
PARTICIPANTS_PAGE_COOKIE = "participants_page"
ENROLLMENTS_PAGE_COOKIE = "enrollments_page"
PAYMENTS_PAGE_COOKIE = "payments_page"


def _get_filters_cookie(request: Request, cookie_name: str) -> dict:
    """Filters are kept out of the URL — read them back from a cookie instead."""
    raw = request.cookies.get(cookie_name)
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else {}
    except (ValueError, TypeError):
        return {}


def _set_filters_cookie_response(
    url: str, cookie_name: str, filters: dict, also_delete: list[str] | None = None
) -> RedirectResponse:
    response = RedirectResponse(url=url, status_code=303)
    if filters:
        response.set_cookie(cookie_name, json.dumps(filters), httponly=True, samesite="lax")
    else:
        response.delete_cookie(cookie_name)
    for extra_cookie in also_delete or []:
        response.delete_cookie(extra_cookie)
    return response


def _sort_redirect(
    url: str,
    cookie_name: str,
    sortable_columns: dict,
    sort_by: str,
    sort_dir: str,
    page_cookie: str | None = None,
) -> RedirectResponse:
    """Validate sort_by/sort_dir and stash them in a cookie instead of the URL."""
    if sort_by not in sortable_columns or sort_dir not in ("asc", "desc"):
        return RedirectResponse(url=url, status_code=303)
    return _set_filters_cookie_response(
        url, cookie_name, {"sort_by": sort_by, "sort_dir": sort_dir}, also_delete=[page_cookie] if page_cookie else None
    )


def _page_size_redirect(
    url: str, cookie_name: str, page_size: str, page_cookie: str | None = None
) -> RedirectResponse:
    """Validate page_size and stash it in a cookie instead of the URL."""
    try:
        size = int(page_size)
    except ValueError:
        return RedirectResponse(url=url, status_code=303)
    if size not in PAGE_SIZE_OPTIONS:
        return RedirectResponse(url=url, status_code=303)
    return _set_filters_cookie_response(
        url, cookie_name, {"page_size": size}, also_delete=[page_cookie] if page_cookie else None
    )


def _page_redirect(url: str, cookie_name: str, page: str) -> RedirectResponse:
    """Validate page and stash it in a cookie instead of the URL."""
    try:
        value = int(page)
    except ValueError:
        return RedirectResponse(url=url, status_code=303)
    if value < 1:
        return RedirectResponse(url=url, status_code=303)
    return _set_filters_cookie_response(url, cookie_name, {"page": value})


def _reservations_redirect(date_iso: str, error: str | None = None) -> RedirectResponse:
    url = "/admin/reservations" + (f"?error={error}" if error else "")
    response = RedirectResponse(url=url, status_code=303)
    response.set_cookie(
        RESERVATIONS_FILTER_COOKIE, json.dumps({"date": date_iso}), httponly=True, samesite="lax"
    )
    return response


templates.env.globals["PAGE_SIZE_OPTIONS"] = PAGE_SIZE_OPTIONS
templates.env.filters["short_date"] = _short_date
templates.env.filters["time12"] = _time12


class NotAuthenticated(Exception):
    pass


@app.exception_handler(NotAuthenticated)
def not_authenticated_handler(request: Request, exc: NotAuthenticated):
    response = RedirectResponse(url=f"/admin/login?next={request.url.path}", status_code=303)
    response.delete_cookie(SESSION_COOKIE)
    return response


def require_admin(request: Request) -> None:
    token = request.cookies.get(SESSION_COOKIE)
    if not token:
        raise NotAuthenticated()
    try:
        get_admin_user(token)
    except Exception:
        raise NotAuthenticated()


admin_router = APIRouter(dependencies=[Depends(require_admin)])

STATUS_BADGE_CLASSES = {
    "pending": "bg-tertiary-fixed/20 text-on-tertiary-fixed-variant",
    "confirmed": "bg-secondary-container/20 text-on-secondary-container",
}

RESERVATION_STATUS_CLASSES = {
    "pending": "bg-tertiary-fixed/20 text-on-tertiary-fixed-variant",
    "confirmed": "bg-secondary/10 text-secondary",
    "cancelled": "bg-error-container/50 text-on-error-container",
}


def format_time_range(starts_at_iso: str, ends_at_iso: str) -> str:
    def fmt(dt: datetime) -> str:
        hour = dt.hour % 12 or 12
        ampm = "AM" if dt.hour < 12 else "PM"
        return f"{hour}:{dt.minute:02d} {ampm}"

    start = datetime.fromisoformat(starts_at_iso)
    end = datetime.fromisoformat(ends_at_iso)
    return f"{fmt(start)} - {fmt(end)}"


def calculate_age(birth_date_str: str) -> int:
    birth_date = date.fromisoformat(birth_date_str)
    today = date.today()
    return today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))


def _next_participant_number() -> str:
    year = date.today().year
    prefix = f"P-{year}-"
    last = (
        supabase.table("participants")
        .select("participant_number")
        .like("participant_number", f"{prefix}%")
        .order("participant_number", desc=True)
        .limit(1)
        .execute()
        .data
    )
    next_seq = int(last[0]["participant_number"].split("-")[-1]) + 1 if last else 1
    return f"{prefix}{next_seq:03d}"


_rate_limit_hits: dict[str, deque] = defaultdict(deque)


def _check_rate_limit(request: Request, bucket: str, max_calls: int = 10, window_seconds: int = 60) -> None:
    """Simple in-memory per-IP throttle for public lookup endpoints. Guards against
    scraping participant PII by enumerating sequential participant numbers."""
    client_ip = request.client.host if request.client else "unknown"
    key = f"{bucket}:{client_ip}"
    now = time.monotonic()
    hits = _rate_limit_hits[key]
    while hits and now - hits[0] > window_seconds:
        hits.popleft()
    if len(hits) >= max_calls:
        raise HTTPException(status_code=429, detail="Too many requests. Please try again shortly.")
    hits.append(now)


@app.get("/")
def index(request: Request):
    return templates.TemplateResponse(request, "index.html")


@app.get("/healthz")
def healthz():
    return {"status": "ok"}


@app.get("/register")
def register_form(request: Request):
    open_cycles = (
        supabase.table("cycles")
        .select("id, name, start_date, end_date")
        .eq("is_open_for_registration", True)
        .order("start_date")
        .execute()
        .data
    )
    return templates.TemplateResponse(request, "register.html", {"open_cycles": open_cycles})


def _normalize_phone(phone: str) -> str:
    """Strip whitespace and add a leading 0 to numbers missing the trunk prefix (e.g. 3644637 -> 03644637)."""
    cleaned = re.sub(r"\s+", "", phone)
    if cleaned.startswith("3"):
        cleaned = f"0{cleaned}"
    return cleaned


@app.get("/register/lookup/participant")
def register_lookup_participant(request: Request, number: str = "", phone: str = ""):
    """Returning-participant shortcut: requires BOTH the participant number and the
    phone number on file to match — the number alone is guessable/sequential, so a
    single-factor lookup would let anyone walk through numbers and harvest another
    family's name/DOB/phone. Rate-limited (tighter than the passive phone lookup)
    since this is effectively a verification check, not just a search."""
    _check_rate_limit(request, "lookup-participant", max_calls=5, window_seconds=300)
    number = number.strip()
    phone = phone.strip()
    if not number or not phone:
        return {"found": False}
    formatted_phone = f"+961{_normalize_phone(phone)}"
    rows = (
        supabase.table("participants")
        .select("id, full_name, mother_name, phone, date_of_birth")
        .ilike("participant_number", number)
        .eq("phone", formatted_phone)
        .limit(1)
        .execute()
        .data
    )
    if not rows:
        return {"found": False}
    participant = rows[0]
    stored_phone = participant["phone"]
    return {
        "found": True,
        "id": participant["id"],
        "full_name": participant["full_name"],
        "mother_name": participant["mother_name"],
        "date_of_birth": participant["date_of_birth"],
        "phone_local": stored_phone[4:] if stored_phone.startswith("+961") else stored_phone,
    }


def _find_exact_participant_match(full_name: str, mother_name: str, formatted_phone: str) -> int | None:
    """Certain match — same as today's implicit dedup. Silent, never needs confirmation."""
    rows = (
        supabase.table("participants")
        .select("id")
        .eq("full_name", full_name)
        .eq("mother_name", mother_name)
        .eq("phone", formatted_phone)
        .execute()
        .data
    )
    return rows[0]["id"] if rows else None


def _find_probable_duplicate(full_name: str, date_of_birth: date) -> dict | None:
    """Silent safety net: exact date-of-birth match + fuzzy name similarity via the
    find_probable_duplicate_participant Postgres function (pg_trgm, threshold 0.35).
    Requires migrations/0008_participant_duplicate_matching.sql to have been run —
    until then the RPC call 404s, which we swallow so registration keeps working
    normally (no confirmation prompts, just no fuzzy safety net) rather than 500ing."""
    try:
        rows = (
            supabase.rpc(
                "find_probable_duplicate_participant",
                {"p_full_name": full_name, "p_date_of_birth": date_of_birth.isoformat()},
            )
            .execute()
            .data
        )
    except APIError:
        return None
    return rows[0] if rows else None


def _upsert_registration(
    full_name: str,
    mother_name: str,
    phone: str,
    date_of_birth: date,
    level: str,
    time_preferred: str,
    cycle_id: int,
    cycle_price: float,
    participant_id_override: int | None = None,
) -> int:
    """Create/update a participant and their enrollment for a cycle. Returns the enrollment id.

    `participant_id_override` is set when the caller already knows which
    participant this is (returning-participant shortcut, phone-match tap, or a
    confirmed "yes" on the duplicate-safety-net modal) — skips matching and
    reuses that participant, refreshing their details from the submitted values.
    """
    formatted_phone = f"+961{_normalize_phone(phone)}"
    if participant_id_override is not None:
        participant_id = participant_id_override
        supabase.table("participants").update({
            "full_name": full_name,
            "mother_name": mother_name,
            "phone": formatted_phone,
            "date_of_birth": date_of_birth.isoformat(),
        }).eq("id", participant_id).execute()
    else:
        existing_id = _find_exact_participant_match(full_name, mother_name, formatted_phone)
        if existing_id is not None:
            participant_id = existing_id
            supabase.table("participants").update(
                {"date_of_birth": date_of_birth.isoformat()}
            ).eq("id", participant_id).execute()
        else:
            participant = (
                supabase.table("participants")
                .insert({
                    "full_name": full_name,
                    "mother_name": mother_name,
                    "phone": formatted_phone,
                    "date_of_birth": date_of_birth.isoformat(),
                    "participant_number": _next_participant_number(),
                })
                .execute()
                .data[0]
            )
            participant_id = participant["id"]

    existing_enrollment = (
        supabase.table("enrollments")
        .select("id")
        .eq("participant_id", participant_id)
        .eq("cycle_id", cycle_id)
        .execute()
        .data
    )
    if existing_enrollment:
        enrollment_id = existing_enrollment[0]["id"]
        supabase.table("enrollments").update(
            {"level": level, "time_preferred": time_preferred}
        ).eq("id", enrollment_id).execute()
    else:
        enrollment = (
            supabase.table("enrollments")
            .insert({
                "participant_id": participant_id,
                "cycle_id": cycle_id,
                "level": level,
                "time_preferred": time_preferred,
                "price": cycle_price,
                "status": "confirmed",
            })
            .execute()
            .data[0]
        )
        enrollment_id = enrollment["id"]

    return enrollment_id


@app.post("/register")
def register_submit(
    request: Request,
    full_name: str = Form(...),
    mother_name: str = Form(...),
    phone: str = Form(...),
    date_of_birth: date = Form(...),
    level: str = Form(...),
    time_preferred: str = Form(...),
    cycle_id: int = Form(...),
    confirmed_participant_id: str = Form(""),
    skip_duplicate_check: str = Form(""),
):
    """Submitted via fetch() from register.js, so this always responds with JSON:
    - {"error": "..."} — validation failure (e.g. cycle closed), form stays open.
    - {"needs_confirmation": True, "match": {...}} — a plausible-but-unconfirmed
      duplicate was found; nothing was created yet, the client shows a modal and
      resubmits with confirmed_participant_id or skip_duplicate_check set.
    - {"redirect_url": "..."} — registration completed, client navigates there.
    """
    open_cycles = (
        supabase.table("cycles")
        .select("id, name, start_date, end_date, cycle_price")
        .eq("is_open_for_registration", True)
        .order("start_date")
        .execute()
        .data
    )
    open_cycle = next((cycle for cycle in open_cycles if cycle["id"] == cycle_id), None)
    if open_cycle is None:
        return JSONResponse(
            {"error": "That cycle is no longer open for registration. Please choose a currently open cycle."},
            status_code=400,
        )

    participant_override = int(confirmed_participant_id) if confirmed_participant_id.strip() else None

    if participant_override is None and not skip_duplicate_check:
        formatted_phone = f"+961{_normalize_phone(phone)}"
        if _find_exact_participant_match(full_name, mother_name, formatted_phone) is None:
            probable_match = _find_probable_duplicate(full_name, date_of_birth)
            if probable_match is not None:
                return JSONResponse({
                    "needs_confirmation": True,
                    "match": {
                        "id": probable_match["id"],
                        "full_name": probable_match["full_name"],
                        "date_of_birth": probable_match["date_of_birth"],
                    },
                })

    enrollment_id = _upsert_registration(
        full_name,
        mother_name,
        phone,
        date_of_birth,
        level,
        time_preferred,
        cycle_id,
        open_cycle["cycle_price"],
        participant_id_override=participant_override,
    )

    return JSONResponse({"redirect_url": f"/register/success/{enrollment_id}"})


@app.get("/register/success/{enrollment_id}")
def register_success(request: Request, enrollment_id: int):
    enrollment = (
        supabase.table("enrollments")
        .select("id, level, time_preferred, status, participants(full_name, participant_number), cycles(name, start_date, end_date)")
        .eq("id", enrollment_id)
        .single()
        .execute()
        .data
    )
    if enrollment is None:
        raise HTTPException(status_code=404)
    return templates.TemplateResponse(request, "register_success.html", {"enrollment": enrollment})


BULK_IMPORT_COLUMNS = ["full_name", "mother_name", "phone", "date_of_birth", "level", "time_preferred", "cycle_id"]
BULK_IMPORT_LEVELS = {"beginner", "intermediate", "advanced"}
BULK_IMPORT_TIME_PREFERRED = {"morning", "afternoon"}


def _clean_phone_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _normalize_phone(str(value).strip())


def _parse_dob_cell(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value).strip())


@app.get("/bulk-import")
def bulk_import_form(request: Request):
    return templates.TemplateResponse(request, "bulk_import.html", {"columns": BULK_IMPORT_COLUMNS, "error": None})


@app.get("/bulk-import/template")
def bulk_import_template():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Registrations"
    worksheet.append(BULK_IMPORT_COLUMNS)
    worksheet.append(["Jane Doe", "Mary Doe", "70123456", "2015-06-01", "beginner", "morning", 1])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bulk_import_template.xlsx"},
    )


@app.post("/bulk-import")
def bulk_import_submit(request: Request, file: UploadFile = File(...)):
    try:
        workbook = load_workbook(filename=io.BytesIO(file.file.read()), data_only=True)
    except Exception:
        return templates.TemplateResponse(
            request,
            "bulk_import.html",
            {
                "columns": BULK_IMPORT_COLUMNS,
                "error": "Could not read that file. Please upload a valid .xlsx file.",
            },
            status_code=400,
        )

    worksheet = workbook.active
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [str(cell).strip().lower() if cell is not None else "" for cell in header_row]
    missing_columns = [column for column in BULK_IMPORT_COLUMNS if column not in headers]
    if missing_columns:
        return templates.TemplateResponse(
            request,
            "bulk_import.html",
            {
                "columns": BULK_IMPORT_COLUMNS,
                "error": f"Missing required column(s): {', '.join(missing_columns)}.",
            },
            status_code=400,
        )

    open_cycles = (
        supabase.table("cycles")
        .select("id, cycle_price")
        .eq("is_open_for_registration", True)
        .execute()
        .data
    )
    open_cycle_price_by_id = {cycle["id"]: cycle["cycle_price"] for cycle in open_cycles}

    results = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(value is None for value in row):
            continue
        row_data = dict(zip(headers, row))
        row_label = str(row_data.get("full_name") or f"Row {row_number}")
        try:
            full_name = str(row_data["full_name"] or "").strip()
            mother_name = str(row_data["mother_name"] or "").strip()
            phone = _clean_phone_cell(row_data["phone"])
            date_of_birth = _parse_dob_cell(row_data["date_of_birth"])
            level = str(row_data["level"] or "").strip().lower()
            time_preferred = str(row_data["time_preferred"] or "").strip().lower()
            cycle_id = int(row_data["cycle_id"])

            if not full_name or not mother_name or not phone:
                raise ValueError("full_name, mother_name, and phone are all required.")
            if level not in BULK_IMPORT_LEVELS:
                raise ValueError(f"level must be one of: {', '.join(sorted(BULK_IMPORT_LEVELS))}.")
            if time_preferred not in BULK_IMPORT_TIME_PREFERRED:
                raise ValueError(f"time_preferred must be one of: {', '.join(sorted(BULK_IMPORT_TIME_PREFERRED))}.")
            if cycle_id not in open_cycle_price_by_id:
                raise ValueError(f"cycle_id {cycle_id} is not currently open for registration.")

            enrollment_id = _upsert_registration(
                full_name,
                mother_name,
                phone,
                date_of_birth,
                level,
                time_preferred,
                cycle_id,
                open_cycle_price_by_id[cycle_id],
            )
            results.append({
                "row": row_number,
                "name": row_label,
                "status": "ok",
                "detail": f"Enrollment #{enrollment_id}",
            })
        except Exception as exc:
            results.append({"row": row_number, "name": row_label, "status": "error", "detail": str(exc)})

    succeeded = sum(1 for result in results if result["status"] == "ok")
    failed = sum(1 for result in results if result["status"] == "error")

    return templates.TemplateResponse(
        request,
        "bulk_import_results.html",
        {"results": results, "succeeded": succeeded, "failed": failed, "total": len(results)},
    )


def _safe_next_path(next_path: str) -> str:
    if next_path.startswith("/") and not next_path.startswith("//"):
        return next_path
    return "/admin/dashboard"


@app.get("/admin/login")
def admin_login_form(request: Request, next: str = "/admin/dashboard"):
    next = _safe_next_path(next)
    if request.cookies.get(SESSION_COOKIE):
        return RedirectResponse(url=next, status_code=303)
    return templates.TemplateResponse(request, "admin/login.html", {"next": next, "error": None})


@app.post("/admin/login")
def admin_login_submit(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    next: str = Form("/admin/dashboard"),
):
    next = _safe_next_path(next)
    try:
        auth_response = sign_in_admin(email, password)
    except Exception:
        return templates.TemplateResponse(
            request,
            "admin/login.html",
            {"next": next, "error": "Invalid email or password."},
            status_code=401,
        )

    response = RedirectResponse(url=next, status_code=303)
    response.set_cookie(
        SESSION_COOKIE,
        auth_response.session.access_token,
        max_age=auth_response.session.expires_in,
        httponly=True,
        samesite="lax",
    )
    return response


@app.post("/admin/logout")
def admin_logout():
    response = RedirectResponse(url="/admin/login", status_code=303)
    response.delete_cookie(SESSION_COOKIE)
    return response


def _get_open_cycle() -> dict | None:
    rows = (
        supabase.table("cycles")
        .select("id, name")
        .eq("is_open_for_registration", True)
        .limit(1)
        .execute()
        .data
    )
    return rows[0] if rows else None


@admin_router.get("/admin/dashboard")
def admin_dashboard(request: Request):
    open_cycle = _get_open_cycle()
    open_cycle_id = open_cycle["id"] if open_cycle else None
    open_cycle_name = open_cycle["name"] if open_cycle else None

    if open_cycle_id is not None:
        enrollments = (
            supabase.table("enrollments")
            .select("id, level, status, cycle_id, time_preferred, created_at, participants(full_name)")
            .eq("cycle_id", open_cycle_id)
            .order("created_at", desc=True)
            .execute()
            .data
        )
    else:
        enrollments = []

    total_reservations = supabase.table("reservations").select("id", count="exact").execute().count
    cancelled_reservations = (
        supabase.table("reservations")
        .select("id", count="exact")
        .eq("status", "cancelled")
        .execute()
        .count
    )

    level_counts = {"beginner": 0, "intermediate": 0, "advanced": 0}
    time_preferred_counts = {"morning": 0, "afternoon": 0}
    for enrollment in enrollments:
        if enrollment["level"] in level_counts:
            level_counts[enrollment["level"]] += 1
        if enrollment["time_preferred"] in time_preferred_counts:
            time_preferred_counts[enrollment["time_preferred"]] += 1
        enrollment["created_at_display"] = datetime.fromisoformat(enrollment["created_at"]).strftime("%B %d, %Y")

    return templates.TemplateResponse(
        request,
        "admin/dashboard.html",
        {
            "today": date.today().strftime("%B %d, %Y"),
            "open_cycle_name": open_cycle_name,
            "total_enrollments": len(enrollments),
            "total_reservations": total_reservations,
            "cancelled_reservations": cancelled_reservations,
            "level_counts": level_counts,
            "time_preferred_counts": time_preferred_counts,
            "recent_enrollments": enrollments[:10],
            "status_classes": STATUS_BADGE_CLASSES,
        },
    )


CYCLES_SORTABLE = {
    "name": "name",
    "start_date": "start_date",
    "end_date": "end_date",
    "cycle_price": "cycle_price",
    "is_open_for_registration": "is_open_for_registration",
}


CYCLE_WEEKDAYS = [
    (0, "Monday"),
    (1, "Tuesday"),
    (2, "Wednesday"),
    (3, "Thursday"),
    (4, "Friday"),
    (5, "Saturday"),
    (6, "Sunday"),
]


@admin_router.get("/admin/cycles")
def admin_cycles(request: Request):
    page, page_size, sort_by, sort_dir = _parse_list_params(
        request,
        CYCLES_SORTABLE,
        default_sort="start_date",
        default_dir="desc",
        sort_cookie=CYCLES_SORT_COOKIE,
        page_size_cookie=CYCLES_PAGE_SIZE_COOKIE,
        page_cookie=CYCLES_PAGE_COOKIE,
    )
    column = CYCLES_SORTABLE[sort_by]

    def build_query(start, end):
        return (
            supabase.table("cycles")
            .select(
                "id, name, start_date, end_date, cycle_price, is_open_for_registration, "
                "pool_id, schedule_days, schedule_start_time, schedule_end_time, pools(name)",
                count="exact",
            )
            .order(column, desc=(sort_dir == "desc"))
            .range(start, end)
            .execute()
        )

    cycles, total, page, total_pages = _fetch_page(build_query, page, page_size)
    pools = supabase.table("pools").select("id, name").order("name").execute().data

    return templates.TemplateResponse(
        request,
        "admin/cycles.html",
        {
            "cycles": cycles,
            "pools": pools,
            "weekdays": CYCLE_WEEKDAYS,
            "error": request.query_params.get("error"),
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": total_pages,
            "sort_by": sort_by,
            "sort_dir": sort_dir,
        },
    )


@admin_router.post("/admin/cycles/sort")
def admin_cycles_set_sort(sort_by: str = Form(...), sort_dir: str = Form(...)):
    return _sort_redirect(
        "/admin/cycles", CYCLES_SORT_COOKIE, CYCLES_SORTABLE, sort_by, sort_dir, page_cookie=CYCLES_PAGE_COOKIE
    )


@admin_router.post("/admin/cycles/page-size")
def admin_cycles_set_page_size(page_size: str = Form(...)):
    return _page_size_redirect(
        "/admin/cycles", CYCLES_PAGE_SIZE_COOKIE, page_size, page_cookie=CYCLES_PAGE_COOKIE
    )


@admin_router.post("/admin/cycles/page")
def admin_cycles_set_page(page: str = Form(...)):
    return _page_redirect("/admin/cycles", CYCLES_PAGE_COOKIE, page)


def _parse_cycle_schedule(
    pool_id: int, schedule_days: list[int], schedule_start_time: str, schedule_end_time: str
) -> dict:
    """Build the cycle-schedule columns, requiring a consistent time window whenever days are picked."""
    days = sorted(set(schedule_days))
    if any(day < 0 or day > 6 for day in days):
        raise ValueError("Invalid day selected.")
    if days and (not schedule_start_time or not schedule_end_time):
        raise ValueError("Pick a start and end time for the cycle's schedule.")
    if days and schedule_end_time <= schedule_start_time:
        raise ValueError("Schedule end time must be after the start time.")
    return {
        "pool_id": pool_id,
        "schedule_days": days,
        "schedule_start_time": schedule_start_time if days else None,
        "schedule_end_time": schedule_end_time if days else None,
    }


@admin_router.post("/admin/cycles")
def admin_cycles_create(
    name: str = Form(...),
    start_date: date = Form(...),
    end_date: date = Form(...),
    cycle_price: float = Form(...),
    pool_id: int = Form(...),
    schedule_days: list[int] = Form([]),
    schedule_start_time: str = Form(""),
    schedule_end_time: str = Form(""),
    is_open_for_registration: bool = Form(False),
):
    try:
        schedule = _parse_cycle_schedule(pool_id, schedule_days, schedule_start_time, schedule_end_time)
    except ValueError as exc:
        return RedirectResponse(url=f"/admin/cycles?error={exc}", status_code=303)
    if is_open_for_registration:
        _close_all_open_cycles()
    try:
        supabase.table("cycles").insert({
            "name": name,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "cycle_price": cycle_price,
            "is_open_for_registration": is_open_for_registration,
            **schedule,
        }).execute()
    except APIError as exc:
        if exc.code == "23P01":
            error = "Those dates overlap an existing cycle. Pick a date range that doesn't conflict."
        elif exc.code == "23514":
            error = "Cycle price must be greater than 0."
        else:
            error = "Could not create the cycle."
        return RedirectResponse(url=f"/admin/cycles?error={error}", status_code=303)
    return RedirectResponse(url="/admin/cycles", status_code=303)


@admin_router.post("/admin/cycles/{cycle_id}/edit")
def admin_cycles_edit(
    cycle_id: int,
    name: str = Form(...),
    start_date: date = Form(...),
    end_date: date = Form(...),
    cycle_price: float = Form(...),
    pool_id: int = Form(...),
    schedule_days: list[int] = Form([]),
    schedule_start_time: str = Form(""),
    schedule_end_time: str = Form(""),
):
    try:
        schedule = _parse_cycle_schedule(pool_id, schedule_days, schedule_start_time, schedule_end_time)
    except ValueError as exc:
        return RedirectResponse(url=f"/admin/cycles?error={exc}", status_code=303)
    try:
        supabase.table("cycles").update({
            "name": name,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "cycle_price": cycle_price,
            **schedule,
        }).eq("id", cycle_id).execute()
    except APIError as exc:
        if exc.code == "23P01":
            error = "Those dates overlap an existing cycle. Pick a date range that doesn't conflict."
        elif exc.code == "23514":
            error = "Cycle price must be greater than 0."
        else:
            error = "Could not save changes."
        return RedirectResponse(url=f"/admin/cycles?error={error}", status_code=303)
    return RedirectResponse(url="/admin/cycles", status_code=303)


def _close_all_open_cycles(except_cycle_id: int | None = None) -> None:
    """Only one cycle may be open for registration at a time."""
    query = (
        supabase.table("cycles")
        .update({"is_open_for_registration": False})
        .eq("is_open_for_registration", True)
    )
    if except_cycle_id is not None:
        query = query.neq("id", except_cycle_id)
    query.execute()


@admin_router.post("/admin/cycles/{cycle_id}/toggle-open")
def admin_cycles_toggle_open(cycle_id: int):
    cycle = (
        supabase.table("cycles")
        .select("is_open_for_registration")
        .eq("id", cycle_id)
        .single()
        .execute()
        .data
    )
    now_open = not cycle["is_open_for_registration"]
    if now_open:
        _close_all_open_cycles(except_cycle_id=cycle_id)
    supabase.table("cycles").update(
        {"is_open_for_registration": now_open}
    ).eq("id", cycle_id).execute()
    return RedirectResponse(url="/admin/cycles", status_code=303)


@admin_router.post("/admin/cycles/{cycle_id}/delete")
def admin_cycles_delete(cycle_id: int):
    try:
        supabase.table("cycles").delete().eq("id", cycle_id).execute()
    except APIError:
        return RedirectResponse(
            url="/admin/cycles?error=Cannot delete a cycle that has enrollments.",
            status_code=303,
        )
    return RedirectResponse(url="/admin/cycles", status_code=303)


PARTICIPANTS_SORTABLE = {
    "full_name": "full_name",
    "participant_number": "participant_number",
    "mother_name": "mother_name",
    "phone": "phone",
    "date_of_birth": "date_of_birth",
}


def _participants_apply_filters(query, filters: dict):
    min_dob, max_dob = _age_bounds_to_dob_range(filters.get("age_min"), filters.get("age_max"))
    if min_dob is not None:
        query = query.gte("date_of_birth", min_dob.isoformat())
    if max_dob is not None:
        query = query.lte("date_of_birth", max_dob.isoformat())
    if filters.get("full_name"):
        query = query.ilike("full_name", f"%{filters['full_name']}%")
    if filters.get("mother_name"):
        query = query.ilike("mother_name", f"%{filters['mother_name']}%")
    if filters.get("phone"):
        query = query.ilike("phone", f"%{filters['phone']}%")
    return query


@admin_router.get("/admin/participants")
def admin_participants(request: Request):
    page, page_size, sort_by, sort_dir = _parse_list_params(
        request,
        PARTICIPANTS_SORTABLE,
        default_sort="full_name",
        default_dir="asc",
        sort_cookie=PARTICIPANTS_SORT_COOKIE,
        page_size_cookie=PARTICIPANTS_PAGE_SIZE_COOKIE,
        page_cookie=PARTICIPANTS_PAGE_COOKIE,
    )
    column = PARTICIPANTS_SORTABLE[sort_by]

    filters = _get_filters_cookie(request, PARTICIPANTS_FILTER_COOKIE)
    full_name = filters.get("full_name")
    mother_name = filters.get("mother_name")
    phone = filters.get("phone")

    def build_query(start, end):
        query = _participants_apply_filters(
            supabase.table("participants").select(
                "id, full_name, mother_name, phone, date_of_birth, participant_number", count="exact"
            ),
            filters,
        )
        return query.order(column, desc=(sort_dir == "desc")).range(start, end).execute()

    participants, total, page, total_pages = _fetch_page(build_query, page, page_size)
    for participant in participants:
        participant["age"] = calculate_age(participant["date_of_birth"])

    return templates.TemplateResponse(
        request,
        "admin/participants.html",
        {
            "participants": participants,
            "error": request.query_params.get("error"),
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": total_pages,
            "sort_by": sort_by,
            "sort_dir": sort_dir,
            "age_min": filters.get("age_min", ""),
            "age_max": filters.get("age_max", ""),
            "full_name_filter": full_name or "",
            "mother_name_filter": mother_name or "",
            "phone_filter": phone or "",
        },
    )


@admin_router.get("/admin/participants/export")
def admin_participants_export(request: Request, format: str = "xlsx"):
    _, _, sort_by, sort_dir = _parse_list_params(
        request,
        PARTICIPANTS_SORTABLE,
        default_sort="full_name",
        default_dir="asc",
        sort_cookie=PARTICIPANTS_SORT_COOKIE,
    )
    column = PARTICIPANTS_SORTABLE[sort_by]

    filters = _get_filters_cookie(request, PARTICIPANTS_FILTER_COOKIE)
    query = _participants_apply_filters(
        supabase.table("participants").select(
            "id, full_name, mother_name, phone, date_of_birth, participant_number"
        ),
        filters,
    )
    participants = query.order(column, desc=(sort_dir == "desc")).execute().data

    headers = ["Participant #", "Full Name", "Mother's Name", "Phone Number", "Date of Birth", "Age"]
    rows = [
        [
            participant["participant_number"],
            participant["full_name"],
            participant["mother_name"],
            participant["phone"],
            participant["date_of_birth"],
            calculate_age(participant["date_of_birth"]),
        ]
        for participant in participants
    ]
    fmt = "csv" if format.lower() == "csv" else "xlsx"
    return _export_response(fmt, "Participants", headers, rows, "participants")


@admin_router.post("/admin/participants/sort")
def admin_participants_set_sort(sort_by: str = Form(...), sort_dir: str = Form(...)):
    return _sort_redirect(
        "/admin/participants",
        PARTICIPANTS_SORT_COOKIE,
        PARTICIPANTS_SORTABLE,
        sort_by,
        sort_dir,
        page_cookie=PARTICIPANTS_PAGE_COOKIE,
    )


@admin_router.post("/admin/participants/page-size")
def admin_participants_set_page_size(page_size: str = Form(...)):
    return _page_size_redirect(
        "/admin/participants", PARTICIPANTS_PAGE_SIZE_COOKIE, page_size, page_cookie=PARTICIPANTS_PAGE_COOKIE
    )


@admin_router.post("/admin/participants/page")
def admin_participants_set_page(page: str = Form(...)):
    return _page_redirect("/admin/participants", PARTICIPANTS_PAGE_COOKIE, page)


@admin_router.post("/admin/participants/filters")
def admin_participants_set_filters(
    age_min: str = Form(""),
    age_max: str = Form(""),
    full_name: str = Form(""),
    mother_name: str = Form(""),
    phone: str = Form(""),
):
    filters = {}
    try:
        if age_min.strip():
            filters["age_min"] = int(age_min)
        if age_max.strip():
            filters["age_max"] = int(age_max)
    except ValueError:
        pass
    if full_name.strip():
        filters["full_name"] = full_name.strip()
    if mother_name.strip():
        filters["mother_name"] = mother_name.strip()
    phone_cleaned = re.sub(r"\s+", "", phone)
    if phone_cleaned:
        filters["phone"] = phone_cleaned

    return _set_filters_cookie_response(
        "/admin/participants", PARTICIPANTS_FILTER_COOKIE, filters, also_delete=[PARTICIPANTS_PAGE_COOKIE]
    )


@admin_router.post("/admin/participants/filters/clear")
def admin_participants_clear_filters():
    return _set_filters_cookie_response(
        "/admin/participants", PARTICIPANTS_FILTER_COOKIE, {}, also_delete=[PARTICIPANTS_PAGE_COOKIE]
    )


@admin_router.post("/admin/participants/{participant_id}/edit")
def admin_participants_edit(
    participant_id: int,
    full_name: str = Form(...),
    mother_name: str = Form(...),
    phone: str = Form(...),
    date_of_birth: date = Form(...),
):
    supabase.table("participants").update({
        "full_name": full_name,
        "mother_name": mother_name,
        "phone": _normalize_phone(phone),
        "date_of_birth": date_of_birth.isoformat(),
    }).eq("id", participant_id).execute()
    return RedirectResponse(url="/admin/participants", status_code=303)


@admin_router.post("/admin/participants/{participant_id}/delete")
def admin_participants_delete(participant_id: int):
    supabase.table("participants").delete().eq("id", participant_id).execute()
    return RedirectResponse(url="/admin/participants", status_code=303)


ENROLLMENTS_SORTABLE = {
    "full_name": ("full_name", "participants"),
    "participant_number": ("participant_number", "participants"),
    "date_of_birth": ("date_of_birth", "participants"),
    "cycle": ("name", "cycles"),
    "level": ("level", None),
    "enrollment_type": ("enrollment_type", None),
    "time_preferred": ("time_preferred", None),
    "start_time": ("start_time", "time_slots"),
    "price": ("price", None),
    "created_at": ("created_at", None),
}

TIME_SLOT_PERIOD_BY_TIME_PREFERRED = {
    "morning": "morning",
    "afternoon": "evening",
}

ENROLLMENTS_SELECT_COLS = (
    "id, level, enrollment_type, time_preferred, price, status, time_slot_id, cycle_id, created_at,"
    " participants!inner(full_name, mother_name, phone, date_of_birth, participant_number),"
    " time_slots(start_time, end_time), cycles(name)"
)


def _resolve_enrollments_cycle_id(filters: dict) -> int | None:
    """The cycle filter defaults to whichever cycle is currently open for
    registration unless the admin has explicitly picked a cycle (or "All
    Cycles", stored as the "all" sentinel) — mirrors the dashboard's default.
    """
    raw = filters.get("cycle_id")
    if raw == "all":
        return None
    if isinstance(raw, int):
        return raw
    open_cycle = _get_open_cycle()
    return open_cycle["id"] if open_cycle else None


def _enrollments_apply_filters(query, filters: dict, cycle_id: int | None = None):
    min_dob, max_dob = _age_bounds_to_dob_range(filters.get("age_min"), filters.get("age_max"))
    if min_dob is not None:
        query = query.gte("participants.date_of_birth", min_dob.isoformat())
    if max_dob is not None:
        query = query.lte("participants.date_of_birth", max_dob.isoformat())
    if cycle_id is not None:
        query = query.eq("cycle_id", cycle_id)
    if filters.get("level"):
        query = query.eq("level", filters["level"])
    if filters.get("enrollment_type"):
        query = query.eq("enrollment_type", filters["enrollment_type"])
    if filters.get("time_preferred"):
        query = query.eq("time_preferred", filters["time_preferred"])
    if filters.get("time_slot_id"):
        query = query.eq("time_slot_id", filters["time_slot_id"])
    if filters.get("full_name"):
        query = query.ilike("participants.full_name", f"%{filters['full_name']}%")
    if filters.get("mother_name"):
        query = query.ilike("participants.mother_name", f"%{filters['mother_name']}%")
    if filters.get("phone"):
        query = query.ilike("participants.phone", f"%{filters['phone']}%")
    return query


def _enrollments_payments_lookup(enrollment_ids: list[int]) -> dict:
    if not enrollment_ids:
        return {}
    payments = (
        supabase.table("payments")
        .select("payable_id, amount")
        .eq("payable_type", "enrollment")
        .in_("payable_id", enrollment_ids)
        .execute()
        .data
    )
    lookup = {}
    for payment in payments:
        lookup[payment["payable_id"]] = lookup.get(payment["payable_id"], 0) + payment["amount"]
    return lookup


def _enrollment_paid_status(enrollment: dict, paid_lookup: dict) -> str:
    paid = paid_lookup.get(enrollment["id"], 0)
    if paid <= 0:
        return "unpaid"
    if paid < enrollment["price"]:
        return "partial"
    return "paid"


def _enrollment_sort_key(column: str, foreign_table: str | None):
    """Sort key for a joined column.

    PostgREST's foreign_table order param only sorts rows *inside* an embedded
    relationship, not the parent rows — so `.order(col, foreign_table=...)` is a
    no-op for sorting enrollments by a participant/time_slot column. Sort those
    columns in Python instead. `value is None` is sorted first in the tuple so
    mixed None/str comparisons (e.g. enrollments with no time slot) never raise.
    """
    def key(enrollment):
        if foreign_table == "participants":
            value = enrollment["participants"].get(column)
        elif foreign_table == "time_slots":
            time_slot = enrollment.get("time_slots")
            value = time_slot.get(column) if time_slot else None
        elif foreign_table == "cycles":
            cycle = enrollment.get("cycles")
            value = cycle.get(column) if cycle else None
        else:
            value = enrollment.get(column)
        return (value is None, value if value is not None else "")

    return key


@admin_router.get("/admin/enrollments")
def admin_enrollments(request: Request):
    page, page_size, sort_by, sort_dir = _parse_list_params(
        request,
        ENROLLMENTS_SORTABLE,
        default_sort="created_at",
        default_dir="desc",
        sort_cookie=ENROLLMENTS_SORT_COOKIE,
        page_size_cookie=ENROLLMENTS_PAGE_SIZE_COOKIE,
        page_cookie=ENROLLMENTS_PAGE_COOKIE,
    )
    column, foreign_table = ENROLLMENTS_SORTABLE[sort_by]

    filters = _get_filters_cookie(request, ENROLLMENTS_FILTER_COOKIE)
    paid_status_filter = filters.get("paid_status")
    cycles = supabase.table("cycles").select("id, name, is_open_for_registration").order("start_date", desc=True).execute().data
    effective_cycle_id = _resolve_enrollments_cycle_id(filters)

    def build_query(start, end):
        query = _enrollments_apply_filters(
            supabase.table("enrollments").select(ENROLLMENTS_SELECT_COLS, count="exact"),
            filters,
            cycle_id=effective_cycle_id,
        )
        return query.order(column, desc=(sort_dir == "desc")).range(start, end).execute()

    if not paid_status_filter and foreign_table is None:
        enrollments, total, page, total_pages = _fetch_page(build_query, page, page_size)
        paid_by_enrollment = _enrollments_payments_lookup([enrollment["id"] for enrollment in enrollments])

        # Totals across every filtered row, not just the current page. Reuses
        # ENROLLMENTS_SELECT_COLS (not just "id, price") because the age filter
        # references participants.date_of_birth, which only resolves when that
        # embed is actually present in the select clause (participants!inner(...)).
        all_matching_rows = (
            _enrollments_apply_filters(
                supabase.table("enrollments").select(ENROLLMENTS_SELECT_COLS), filters, cycle_id=effective_cycle_id
            )
            .execute()
            .data
        )
        total_amount = sum(row["price"] for row in all_matching_rows)
        all_matching_paid = _enrollments_payments_lookup([row["id"] for row in all_matching_rows])
        total_paid = sum(all_matching_paid.values())
    else:
        # Two cases land here: paid_status is derived from the payments table (not
        # a real column, so it can't be pushed into the query), and sorting by a
        # joined participants/time_slots column (PostgREST's foreign_table order
        # only reorders rows *inside* an embedded relationship, not the parent
        # rows). Both require fetching everything matching the other filters,
        # then filtering/sorting/paginating in Python.
        query = _enrollments_apply_filters(
            supabase.table("enrollments").select(ENROLLMENTS_SELECT_COLS), filters, cycle_id=effective_cycle_id
        )
        all_matching = query.execute().data
        paid_lookup = _enrollments_payments_lookup([enrollment["id"] for enrollment in all_matching])

        filtered = (
            [
                enrollment
                for enrollment in all_matching
                if _enrollment_paid_status(enrollment, paid_lookup) == paid_status_filter
            ]
            if paid_status_filter
            else all_matching
        )
        filtered.sort(key=_enrollment_sort_key(column, foreign_table), reverse=(sort_dir == "desc"))

        total = len(filtered)
        total_pages = max(1, math.ceil(total / page_size)) if total else 1
        page = min(max(page, 1), total_pages)
        start = (page - 1) * page_size
        enrollments = filtered[start : start + page_size]
        paid_by_enrollment = paid_lookup

        # Totals across every filtered row, not just the current page.
        total_amount = sum(enrollment["price"] for enrollment in filtered)
        total_paid = sum(paid_lookup.get(enrollment["id"], 0) for enrollment in filtered)

    time_slots = supabase.table("time_slots").select("id, start_time, end_time, period").order("start_time").execute().data

    for enrollment in enrollments:
        enrollment["age"] = calculate_age(enrollment["participants"]["date_of_birth"])
        enrollment["paid"] = paid_by_enrollment.get(enrollment["id"], 0)
        enrollment["remaining"] = enrollment["price"] - enrollment["paid"]

    return templates.TemplateResponse(
        request,
        "admin/enrollments.html",
        {
            "enrollments": enrollments,
            "time_slots": time_slots,
            "cycles": cycles,
            "selected_cycle_id": effective_cycle_id,
            "status_classes": STATUS_BADGE_CLASSES,
            "today_iso": date.today().isoformat(),
            "error": request.query_params.get("error"),
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": total_pages,
            "sort_by": sort_by,
            "sort_dir": sort_dir,
            "filters": filters,
            "total_amount": total_amount,
            "total_paid": total_paid,
        },
    )


@admin_router.get("/admin/enrollments/export")
def admin_enrollments_export(request: Request, format: str = "xlsx"):
    _, _, sort_by, sort_dir = _parse_list_params(
        request,
        ENROLLMENTS_SORTABLE,
        default_sort="created_at",
        default_dir="desc",
        sort_cookie=ENROLLMENTS_SORT_COOKIE,
    )
    column, foreign_table = ENROLLMENTS_SORTABLE[sort_by]

    filters = _get_filters_cookie(request, ENROLLMENTS_FILTER_COOKIE)
    paid_status_filter = filters.get("paid_status")
    effective_cycle_id = _resolve_enrollments_cycle_id(filters)

    query = _enrollments_apply_filters(
        supabase.table("enrollments").select(ENROLLMENTS_SELECT_COLS), filters, cycle_id=effective_cycle_id
    )
    enrollments = query.execute().data

    paid_lookup = _enrollments_payments_lookup([enrollment["id"] for enrollment in enrollments])
    if paid_status_filter:
        enrollments = [
            enrollment
            for enrollment in enrollments
            if _enrollment_paid_status(enrollment, paid_lookup) == paid_status_filter
        ]
    enrollments.sort(key=_enrollment_sort_key(column, foreign_table), reverse=(sort_dir == "desc"))

    headers = [
        "Participant #",
        "Name",
        "Mother's Name",
        "Phone Number",
        "Age",
        "Cycle",
        "Level",
        "Type",
        "Preferred Time",
        "Time Slot",
        "Amount",
        "Paid",
        "Remaining",
        "Status",
    ]
    rows = []
    for enrollment in enrollments:
        participant = enrollment["participants"]
        cycle = enrollment.get("cycles")
        paid = paid_lookup.get(enrollment["id"], 0)
        time_slot = enrollment.get("time_slots")
        time_slot_label = f"{_time12(time_slot['start_time'])}–{_time12(time_slot['end_time'])}" if time_slot else ""
        rows.append(
            [
                participant["participant_number"],
                participant["full_name"],
                participant["mother_name"],
                participant["phone"],
                calculate_age(participant["date_of_birth"]),
                cycle["name"] if cycle else "",
                enrollment["level"].capitalize(),
                enrollment["enrollment_type"].capitalize(),
                enrollment["time_preferred"].capitalize(),
                time_slot_label,
                enrollment["price"],
                paid,
                enrollment["price"] - paid,
                enrollment["status"].capitalize(),
            ]
        )
    fmt = "csv" if format.lower() == "csv" else "xlsx"
    return _export_response(fmt, "Enrollments", headers, rows, "enrollments")


@admin_router.post("/admin/enrollments/sort")
def admin_enrollments_set_sort(sort_by: str = Form(...), sort_dir: str = Form(...)):
    return _sort_redirect(
        "/admin/enrollments",
        ENROLLMENTS_SORT_COOKIE,
        ENROLLMENTS_SORTABLE,
        sort_by,
        sort_dir,
        page_cookie=ENROLLMENTS_PAGE_COOKIE,
    )


@admin_router.post("/admin/enrollments/page-size")
def admin_enrollments_set_page_size(page_size: str = Form(...)):
    return _page_size_redirect(
        "/admin/enrollments", ENROLLMENTS_PAGE_SIZE_COOKIE, page_size, page_cookie=ENROLLMENTS_PAGE_COOKIE
    )


@admin_router.post("/admin/enrollments/page")
def admin_enrollments_set_page(page: str = Form(...)):
    return _page_redirect("/admin/enrollments", ENROLLMENTS_PAGE_COOKIE, page)


@admin_router.post("/admin/enrollments/filters")
def admin_enrollments_set_filters(
    cycle_id: str = Form(""),
    age_min: str = Form(""),
    age_max: str = Form(""),
    level: str = Form(""),
    enrollment_type: str = Form(""),
    time_preferred: str = Form(""),
    time_slot_id: str = Form(""),
    paid_status: str = Form(""),
    full_name: str = Form(""),
    mother_name: str = Form(""),
    phone: str = Form(""),
):
    filters = {}
    if cycle_id == "all":
        filters["cycle_id"] = "all"
    try:
        if age_min.strip():
            filters["age_min"] = int(age_min)
        if age_max.strip():
            filters["age_max"] = int(age_max)
        if time_slot_id.strip():
            filters["time_slot_id"] = int(time_slot_id)
        if cycle_id.strip() and cycle_id != "all":
            filters["cycle_id"] = int(cycle_id)
    except ValueError:
        pass
    if level:
        filters["level"] = level
    if enrollment_type:
        filters["enrollment_type"] = enrollment_type
    if time_preferred:
        filters["time_preferred"] = time_preferred
    if paid_status:
        filters["paid_status"] = paid_status
    if full_name.strip():
        filters["full_name"] = full_name.strip()
    if mother_name.strip():
        filters["mother_name"] = mother_name.strip()
    phone_cleaned = re.sub(r"\s+", "", phone)
    if phone_cleaned:
        filters["phone"] = phone_cleaned

    return _set_filters_cookie_response(
        "/admin/enrollments", ENROLLMENTS_FILTER_COOKIE, filters, also_delete=[ENROLLMENTS_PAGE_COOKIE]
    )


@admin_router.post("/admin/enrollments/filters/clear")
def admin_enrollments_clear_filters():
    return _set_filters_cookie_response(
        "/admin/enrollments", ENROLLMENTS_FILTER_COOKIE, {}, also_delete=[ENROLLMENTS_PAGE_COOKIE]
    )


@admin_router.post("/admin/enrollments/{enrollment_id}/edit")
def admin_enrollments_edit(
    enrollment_id: int,
    time_preferred: str = Form(...),
    level: str = Form(...),
    enrollment_type: str = Form(...),
    status: str = Form(...),
    price: float = Form(...),
    time_slot_id: str = Form(""),
):
    try:
        supabase.table("enrollments").update({
            "time_preferred": time_preferred,
            "level": level,
            "enrollment_type": enrollment_type,
            "status": status,
            "price": price,
            "time_slot_id": int(time_slot_id) if time_slot_id else None,
        }).eq("id", enrollment_id).execute()
    except APIError:
        return RedirectResponse(
            url="/admin/enrollments?error=Could not save changes. The status value may not be recognized by the database.",
            status_code=303,
        )
    return RedirectResponse(url="/admin/enrollments", status_code=303)


@admin_router.post("/admin/enrollments/{enrollment_id}/record-payment")
def admin_enrollments_record_payment(
    enrollment_id: int,
    amount: float = Form(...),
    paid_at: date = Form(...),
    method: str = Form(...),
    notes: str = Form(""),
):
    enrollment = (
        supabase.table("enrollments")
        .select("price")
        .eq("id", enrollment_id)
        .single()
        .execute()
        .data
    )
    payments = (
        supabase.table("payments")
        .select("amount")
        .eq("payable_type", "enrollment")
        .eq("payable_id", enrollment_id)
        .execute()
        .data
    )
    paid_so_far = sum(payment["amount"] for payment in payments)
    remaining = enrollment["price"] - paid_so_far
    if amount > remaining:
        return RedirectResponse(
            url=f"/admin/enrollments?error=Payment of ${amount:.2f} exceeds the remaining balance of ${remaining:.2f}.",
            status_code=303,
        )

    supabase.table("payments").insert({
        "payable_type": "enrollment",
        "payable_id": enrollment_id,
        "amount": amount,
        "method": method,
        "paid_at": paid_at.isoformat(),
        "notes": notes or None,
    }).execute()
    return RedirectResponse(url="/admin/enrollments", status_code=303)


@admin_router.post("/admin/enrollments/{enrollment_id}/delete")
def admin_enrollments_delete(enrollment_id: int):
    supabase.table("enrollments").delete().eq("id", enrollment_id).execute()
    return RedirectResponse(url="/admin/enrollments", status_code=303)


def _parse_reservation_window(
    start_date: date_cls, start_time: str, end_date: date_cls, end_time: str
) -> tuple[str, str]:
    """Combine separate date/time inputs into ISO datetimes, requiring end strictly after start."""
    starts_at = f"{start_date.isoformat()}T{start_time}:00"
    ends_at = f"{end_date.isoformat()}T{end_time}:00"
    if ends_at <= starts_at:
        raise ValueError("End date/time must be after the start date/time.")
    return starts_at, ends_at


def _resolve_base_price(
    pool_id: int, start_date: date_cls, end_date: date_cls, is_evening_stay: bool
) -> float:
    """Pick the pool's price tier for a reservation.

    The day/night rate is keyed off the checkout date when the stay spans multiple
    calendar days (e.g. Friday -> Saturday uses the weekend rate), since that's the
    day the guest actually wakes up on. Evening stay overrides the tier entirely.
    """
    is_night_stay = end_date != start_date
    rate_date = end_date if is_night_stay else start_date
    day_type = "weekend" if rate_date.weekday() >= 5 else "weekday"
    pricing_rule = (
        supabase.table("pricing_rules")
        .select("price, night_price, evening_price")
        .eq("pool_id", pool_id)
        .eq("day_type", day_type)
        .limit(1)
        .execute()
        .data
    )
    if not pricing_rule:
        return 0
    if is_evening_stay:
        return pricing_rule[0]["evening_price"]
    if is_night_stay:
        return pricing_rule[0]["night_price"]
    return pricing_rule[0]["price"]


def _find_conflicting_cycle(pool_id: int, starts_at: str, ends_at: str) -> dict | None:
    """Find an active recurring cycle on this pool whose scheduled days/time overlap the reservation window."""
    start_dt = datetime.fromisoformat(starts_at)
    end_dt = datetime.fromisoformat(ends_at)
    cycles = (
        supabase.table("cycles")
        .select("id, name, schedule_days, schedule_start_time, schedule_end_time")
        .eq("pool_id", pool_id)
        .lte("start_date", end_dt.date().isoformat())
        .gte("end_date", start_dt.date().isoformat())
        .execute()
        .data
    )
    day = start_dt.date()
    while day <= end_dt.date():
        for cycle in cycles:
            if day.weekday() not in cycle["schedule_days"]:
                continue
            if not cycle["schedule_start_time"] or not cycle["schedule_end_time"]:
                continue
            cycle_start = datetime.fromisoformat(f"{day.isoformat()}T{cycle['schedule_start_time']}")
            cycle_end = datetime.fromisoformat(f"{day.isoformat()}T{cycle['schedule_end_time']}")
            if start_dt < cycle_end and end_dt > cycle_start:
                return cycle
        day += timedelta(days=1)
    return None


@admin_router.get("/admin/reservations")
def admin_reservations(request: Request):
    pools = supabase.table("pools").select("id, name").order("name").execute().data

    filters = _get_filters_cookie(request, RESERVATIONS_FILTER_COOKIE)
    try:
        selected_date = date_cls.fromisoformat(filters.get("date", ""))
    except (ValueError, TypeError):
        selected_date = date_cls.today()

    if not pools:
        return templates.TemplateResponse(
            request,
            "admin/reservations.html",
            {"pools": pools, "error": request.query_params.get("error")},
        )

    day_start = datetime.combine(selected_date, datetime.min.time()).isoformat()
    day_end = datetime.combine(selected_date + timedelta(days=1), datetime.min.time()).isoformat()

    # A reservation shows on this day if its window overlaps the day at all —
    # not just when it starts here — so overnight bookings (e.g. 11 PM to
    # noon the next day) also appear on the day they run into.
    reservations = (
        supabase.table("reservations")
        .select(
            "id, pool_id, customer_name, customer_phone, starts_at, ends_at, "
            "price_snapshot, base_price, extra_charge_applied, extra_charge_amount, "
            "is_evening_stay, status"
        )
        .lt("starts_at", day_end)
        .gt("ends_at", day_start)
        .order("starts_at")
        .execute()
        .data
    )

    reservation_ids = [reservation["id"] for reservation in reservations]
    paid_by_reservation = {}
    if reservation_ids:
        payments = (
            supabase.table("payments")
            .select("payable_id, amount")
            .eq("payable_type", "reservation")
            .in_("payable_id", reservation_ids)
            .execute()
            .data
        )
        for payment in payments:
            paid_by_reservation[payment["payable_id"]] = (
                paid_by_reservation.get(payment["payable_id"], 0) + payment["amount"]
            )

    reservations_by_pool: dict[int, list] = {}
    for reservation in reservations:
        reservation["is_cycle"] = False
        reservation["time_range"] = format_time_range(reservation["starts_at"], reservation["ends_at"])
        reservation["paid"] = paid_by_reservation.get(reservation["id"], 0)
        reservation["remaining"] = reservation["price_snapshot"] - reservation["paid"]
        starts_dt = datetime.fromisoformat(reservation["starts_at"])
        ends_dt = datetime.fromisoformat(reservation["ends_at"])
        reservation["start_date"] = starts_dt.date().isoformat()
        reservation["start_time"] = starts_dt.strftime("%H:%M")
        reservation["end_date"] = ends_dt.date().isoformat()
        reservation["end_time"] = ends_dt.strftime("%H:%M")
        reservation["phone_local"] = reservation["customer_phone"].removeprefix("+961").lstrip()
        reservations_by_pool.setdefault(reservation["pool_id"], []).append(reservation)

    # Cycles with a matching recurring schedule show up as read-only blocks
    # alongside real reservations for their pool on days they run.
    cycles_today = (
        supabase.table("cycles")
        .select("id, name, pool_id, schedule_start_time, schedule_end_time")
        .lte("start_date", selected_date.isoformat())
        .gte("end_date", selected_date.isoformat())
        .contains("schedule_days", [str(selected_date.weekday())])
        .execute()
        .data
    )
    for cycle in cycles_today:
        if cycle["pool_id"] is None or not cycle["schedule_start_time"] or not cycle["schedule_end_time"]:
            continue
        start_time = cycle["schedule_start_time"][:5]
        end_time = cycle["schedule_end_time"][:5]
        reservations_by_pool.setdefault(cycle["pool_id"], []).append({
            "is_cycle": True,
            "cycle_name": cycle["name"],
            "start_time": start_time,
            "end_time": end_time,
            "time_range": format_time_range(
                f"{selected_date.isoformat()}T{start_time}:00",
                f"{selected_date.isoformat()}T{end_time}:00",
            ),
        })

    for pool_reservations in reservations_by_pool.values():
        pool_reservations.sort(key=lambda entry: entry["start_time"])

    return templates.TemplateResponse(
        request,
        "admin/reservations.html",
        {
            "pools": pools,
            "reservations_by_pool": reservations_by_pool,
            "status_classes": RESERVATION_STATUS_CLASSES,
            "display_date": selected_date.strftime("%B %d, %Y"),
            "weekday_name": selected_date.strftime("%A"),
            "iso_date": selected_date.isoformat(),
            "prev_date": (selected_date - timedelta(days=1)).isoformat(),
            "next_date": (selected_date + timedelta(days=1)).isoformat(),
            "today_iso": date_cls.today().isoformat(),
            "error": request.query_params.get("error"),
        },
    )


@admin_router.post("/admin/reservations/filters")
def admin_reservations_set_date(date: str = Form("")):
    try:
        date_iso = date_cls.fromisoformat(date).isoformat()
    except (ValueError, TypeError):
        return _set_filters_cookie_response("/admin/reservations", RESERVATIONS_FILTER_COOKIE, {})
    return _reservations_redirect(date_iso)


@admin_router.post("/admin/reservations")
def admin_reservations_create(
    pool_id: int = Form(...),
    customer_name: str = Form(...),
    customer_phone: str = Form(...),
    start_date: date_cls = Form(...),
    start_time: str = Form(...),
    end_date: date_cls = Form(...),
    end_time: str = Form(...),
    status: str = Form("pending"),
    extra_charge_amount: float = Form(0),
    is_evening_stay: bool = Form(False),
):
    try:
        starts_at, ends_at = _parse_reservation_window(start_date, start_time, end_date, end_time)
    except ValueError as exc:
        return _reservations_redirect(start_date.isoformat(), str(exc))

    conflicting_cycle = _find_conflicting_cycle(pool_id, starts_at, ends_at)
    if conflicting_cycle:
        return _reservations_redirect(
            start_date.isoformat(),
            f"This pool is booked for the cycle \"{conflicting_cycle['name']}\" during that time. Pick a different time or pool.",
        )

    base_price = _resolve_base_price(pool_id, start_date, end_date, is_evening_stay)
    price = base_price + extra_charge_amount

    try:
        supabase.table("reservations").insert({
            "pool_id": pool_id,
            "customer_name": customer_name,
            "customer_phone": f"+961{_normalize_phone(customer_phone)}",
            "starts_at": starts_at,
            "ends_at": ends_at,
            "price_snapshot": price,
            "base_price": base_price,
            "extra_charge_applied": extra_charge_amount > 0,
            "extra_charge_amount": extra_charge_amount,
            "is_evening_stay": is_evening_stay,
            "status": status,
        }).execute()
    except APIError as exc:
        if exc.code == "23P01":
            error = "That pool is already booked during part of this time range. Pick a different time or pool."
        else:
            error = "Could not create reservation. Check the status value."
        return _reservations_redirect(start_date.isoformat(), error)
    return _reservations_redirect(start_date.isoformat())


@admin_router.post("/admin/reservations/{reservation_id}/edit")
def admin_reservations_edit(
    reservation_id: int,
    pool_id: int = Form(...),
    customer_name: str = Form(...),
    customer_phone: str = Form(...),
    start_date: date_cls = Form(...),
    start_time: str = Form(...),
    end_date: date_cls = Form(...),
    end_time: str = Form(...),
    status: str = Form(...),
    extra_charge_amount: float = Form(0),
    is_evening_stay: bool = Form(False),
):
    try:
        starts_at, ends_at = _parse_reservation_window(start_date, start_time, end_date, end_time)
    except ValueError as exc:
        return _reservations_redirect(start_date.isoformat(), str(exc))

    conflicting_cycle = _find_conflicting_cycle(pool_id, starts_at, ends_at)
    if conflicting_cycle:
        return _reservations_redirect(
            start_date.isoformat(),
            f"This pool is booked for the cycle \"{conflicting_cycle['name']}\" during that time. Pick a different time or pool.",
        )

    base_price = _resolve_base_price(pool_id, start_date, end_date, is_evening_stay)
    price = base_price + extra_charge_amount

    try:
        supabase.table("reservations").update({
            "pool_id": pool_id,
            "customer_name": customer_name,
            "customer_phone": f"+961{_normalize_phone(customer_phone)}",
            "starts_at": starts_at,
            "ends_at": ends_at,
            "price_snapshot": price,
            "base_price": base_price,
            "extra_charge_applied": extra_charge_amount > 0,
            "extra_charge_amount": extra_charge_amount,
            "is_evening_stay": is_evening_stay,
            "status": status,
        }).eq("id", reservation_id).execute()
    except APIError as exc:
        if exc.code == "23P01":
            error = "That pool is already booked during part of this time range. Pick a different time or pool."
        else:
            error = "Could not save changes. Check the status value."
        return _reservations_redirect(start_date.isoformat(), error)
    return _reservations_redirect(start_date.isoformat())


@admin_router.post("/admin/reservations/{reservation_id}/record-payment")
def admin_reservations_record_payment(
    reservation_id: int,
    amount: float = Form(...),
    paid_at: date_cls = Form(...),
    method: str = Form(...),
    notes: str = Form(""),
):
    reservation = (
        supabase.table("reservations")
        .select("price_snapshot, starts_at")
        .eq("id", reservation_id)
        .single()
        .execute()
        .data
    )
    reservation_date = datetime.fromisoformat(reservation["starts_at"]).date().isoformat()

    payments = (
        supabase.table("payments")
        .select("amount")
        .eq("payable_type", "reservation")
        .eq("payable_id", reservation_id)
        .execute()
        .data
    )
    paid_so_far = sum(payment["amount"] for payment in payments)
    remaining = reservation["price_snapshot"] - paid_so_far
    if amount > remaining:
        return _reservations_redirect(
            reservation_date,
            f"Payment of ${amount:.2f} exceeds the remaining balance of ${remaining:.2f}.",
        )

    supabase.table("payments").insert({
        "payable_type": "reservation",
        "payable_id": reservation_id,
        "amount": amount,
        "method": method,
        "paid_at": paid_at.isoformat(),
        "notes": notes or None,
    }).execute()
    return _reservations_redirect(reservation_date)


@admin_router.post("/admin/reservations/{reservation_id}/cancel")
def admin_reservations_cancel(reservation_id: int):
    reservation = (
        supabase.table("reservations")
        .select("starts_at")
        .eq("id", reservation_id)
        .single()
        .execute()
        .data
    )
    reservation_date = datetime.fromisoformat(reservation["starts_at"]).date().isoformat()
    supabase.table("reservations").update({"status": "cancelled"}).eq("id", reservation_id).execute()
    return _reservations_redirect(reservation_date)


PAYMENTS_SORTABLE = {
    "payable_type": "payable_type",
    "amount": "amount",
    "method": "method",
    "paid_at": "paid_at",
}
PAYMENT_METHODS = ["cash", "whish"]
PAYMENT_SOURCES = ["enrollment", "reservation"]


@admin_router.get("/admin/payments")
def admin_payments(request: Request):
    page, page_size, sort_by, sort_dir = _parse_list_params(
        request,
        PAYMENTS_SORTABLE,
        default_sort="paid_at",
        default_dir="desc",
        sort_cookie=PAYMENTS_SORT_COOKIE,
        page_size_cookie=PAYMENTS_PAGE_SIZE_COOKIE,
        page_cookie=PAYMENTS_PAGE_COOKIE,
    )
    column = PAYMENTS_SORTABLE[sort_by]

    filters = _get_filters_cookie(request, PAYMENTS_FILTER_COOKIE)
    date_from = filters.get("date_from")
    date_to = filters.get("date_to")
    method = filters.get("method")
    source = filters.get("source")

    try:
        date_from = date_cls.fromisoformat(date_from).isoformat() if date_from else None
    except ValueError:
        date_from = None
    try:
        date_to = date_cls.fromisoformat(date_to).isoformat() if date_to else None
    except ValueError:
        date_to = None
    if method not in PAYMENT_METHODS:
        method = None
    if source not in PAYMENT_SOURCES:
        source = None

    def apply_filters(query):
        if date_from:
            query = query.gte("paid_at", f"{date_from}T00:00:00")
        if date_to:
            query = query.lte("paid_at", f"{date_to}T23:59:59")
        if method:
            query = query.eq("method", method)
        if source:
            query = query.eq("payable_type", source)
        return query

    def build_query(start, end):
        query = apply_filters(
            supabase.table("payments").select("id, payable_type, payable_id, amount, method, paid_at", count="exact")
        )
        return query.order(column, desc=(sort_dir == "desc")).range(start, end).execute()

    payments, total, page, total_pages = _fetch_page(build_query, page, page_size)

    # Sum across every filtered row, not just the current page.
    filtered_amounts = apply_filters(supabase.table("payments").select("amount")).execute().data
    total_amount = sum(row["amount"] for row in filtered_amounts)

    enrollment_ids = [p["payable_id"] for p in payments if p["payable_type"] == "enrollment"]
    reservation_ids = [p["payable_id"] for p in payments if p["payable_type"] == "reservation"]

    name_by_enrollment_id = {}
    if enrollment_ids:
        rows = (
            supabase.table("enrollments")
            .select("id, participants(full_name)")
            .in_("id", enrollment_ids)
            .execute()
            .data
        )
        for row in rows:
            name_by_enrollment_id[row["id"]] = row["participants"]["full_name"] if row["participants"] else None

    name_by_reservation_id = {}
    if reservation_ids:
        rows = (
            supabase.table("reservations")
            .select("id, customer_name")
            .in_("id", reservation_ids)
            .execute()
            .data
        )
        for row in rows:
            name_by_reservation_id[row["id"]] = row["customer_name"]

    for payment in payments:
        if payment["payable_type"] == "enrollment":
            payment["source_label"] = "Enrollment"
            payment["source_name"] = name_by_enrollment_id.get(payment["payable_id"])
        else:
            payment["source_label"] = "Reservation"
            payment["source_name"] = name_by_reservation_id.get(payment["payable_id"])

    return templates.TemplateResponse(
        request,
        "admin/payments.html",
        {
            "payments": payments,
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": total_pages,
            "sort_by": sort_by,
            "sort_dir": sort_dir,
            "date_from": date_from or "",
            "date_to": date_to or "",
            "method_filter": method or "",
            "payment_methods": PAYMENT_METHODS,
            "source_filter": source or "",
            "payment_sources": PAYMENT_SOURCES,
            "total_amount": total_amount,
        },
    )


@admin_router.post("/admin/payments/sort")
def admin_payments_set_sort(sort_by: str = Form(...), sort_dir: str = Form(...)):
    return _sort_redirect(
        "/admin/payments",
        PAYMENTS_SORT_COOKIE,
        PAYMENTS_SORTABLE,
        sort_by,
        sort_dir,
        page_cookie=PAYMENTS_PAGE_COOKIE,
    )


@admin_router.post("/admin/payments/page-size")
def admin_payments_set_page_size(page_size: str = Form(...)):
    return _page_size_redirect(
        "/admin/payments", PAYMENTS_PAGE_SIZE_COOKIE, page_size, page_cookie=PAYMENTS_PAGE_COOKIE
    )


@admin_router.post("/admin/payments/page")
def admin_payments_set_page(page: str = Form(...)):
    return _page_redirect("/admin/payments", PAYMENTS_PAGE_COOKIE, page)


@admin_router.post("/admin/payments/filters")
def admin_payments_set_filters(
    date_from: str = Form(""),
    date_to: str = Form(""),
    method: str = Form(""),
    source: str = Form(""),
):
    filters = {}
    try:
        if date_from.strip():
            date_cls.fromisoformat(date_from.strip())
            filters["date_from"] = date_from.strip()
    except ValueError:
        pass
    try:
        if date_to.strip():
            date_cls.fromisoformat(date_to.strip())
            filters["date_to"] = date_to.strip()
    except ValueError:
        pass
    if method in PAYMENT_METHODS:
        filters["method"] = method
    if source in PAYMENT_SOURCES:
        filters["source"] = source

    return _set_filters_cookie_response(
        "/admin/payments", PAYMENTS_FILTER_COOKIE, filters, also_delete=[PAYMENTS_PAGE_COOKIE]
    )


@admin_router.post("/admin/payments/filters/clear")
def admin_payments_clear_filters():
    return _set_filters_cookie_response(
        "/admin/payments", PAYMENTS_FILTER_COOKIE, {}, also_delete=[PAYMENTS_PAGE_COOKIE]
    )


@admin_router.post("/admin/payments/{payment_id}/delete")
def admin_payments_delete(payment_id: int):
    supabase.table("payments").delete().eq("id", payment_id).execute()
    return RedirectResponse(url="/admin/payments", status_code=303)


@admin_router.get("/admin/settings")
def admin_settings(request: Request):
    pools = supabase.table("pools").select("id, name, capacity, is_active").order("name").execute().data
    pricing_rules = (
        supabase.table("pricing_rules")
        .select("pool_id, day_type, price, night_price, evening_price")
        .execute()
        .data
    )

    rules_by_pool: dict[int, dict[str, dict]] = {}
    for rule in pricing_rules:
        rules_by_pool.setdefault(rule["pool_id"], {})[rule["day_type"]] = rule

    for pool in pools:
        weekday_rule = rules_by_pool.get(pool["id"], {}).get("weekday")
        weekend_rule = rules_by_pool.get(pool["id"], {}).get("weekend")
        pool["weekday_price"] = weekday_rule["price"] if weekday_rule else None
        pool["weekday_night_price"] = weekday_rule["night_price"] if weekday_rule else 0
        pool["weekday_evening_price"] = weekday_rule["evening_price"] if weekday_rule else 0
        pool["weekend_price"] = weekend_rule["price"] if weekend_rule else None
        pool["weekend_night_price"] = weekend_rule["night_price"] if weekend_rule else 0
        pool["weekend_evening_price"] = weekend_rule["evening_price"] if weekend_rule else 0

    return templates.TemplateResponse(
        request,
        "admin/settings.html",
        {"pools": pools, "error": request.query_params.get("error")},
    )


@admin_router.post("/admin/settings/pools")
def admin_settings_create_pool(name: str = Form(...), capacity: int = Form(...)):
    supabase.table("pools").insert({"name": name, "capacity": capacity}).execute()
    return RedirectResponse(url="/admin/settings", status_code=303)


@admin_router.post("/admin/settings/pools/{pool_id}/toggle-active")
def admin_settings_toggle_pool_active(pool_id: int):
    pool = supabase.table("pools").select("is_active").eq("id", pool_id).single().execute().data
    supabase.table("pools").update({"is_active": not pool["is_active"]}).eq("id", pool_id).execute()
    return RedirectResponse(url="/admin/settings", status_code=303)


def _upsert_pricing_rule(
    pool_id: int, day_type: str, price: float, night_price: float, evening_price: float
) -> None:
    existing = (
        supabase.table("pricing_rules")
        .select("id")
        .eq("pool_id", pool_id)
        .eq("day_type", day_type)
        .execute()
        .data
    )
    if existing:
        supabase.table("pricing_rules").update(
            {"price": price, "night_price": night_price, "evening_price": evening_price}
        ).eq("id", existing[0]["id"]).execute()
    else:
        supabase.table("pricing_rules").insert(
            {
                "pool_id": pool_id,
                "day_type": day_type,
                "price": price,
                "night_price": night_price,
                "evening_price": evening_price,
            }
        ).execute()


@admin_router.post("/admin/settings/pools/{pool_id}/pricing")
def admin_settings_update_pricing(
    pool_id: int,
    weekday_price: float = Form(...),
    weekend_price: float = Form(...),
    weekday_night_price: float = Form(0),
    weekend_night_price: float = Form(0),
    weekday_evening_price: float = Form(0),
    weekend_evening_price: float = Form(0),
):
    _upsert_pricing_rule(pool_id, "weekday", weekday_price, weekday_night_price, weekday_evening_price)
    _upsert_pricing_rule(pool_id, "weekend", weekend_price, weekend_night_price, weekend_evening_price)
    return RedirectResponse(url="/admin/settings", status_code=303)


app.include_router(admin_router)
